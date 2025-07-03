
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re

# Clasificar tipo de vigencia
def clasificar_vigencia(inicio_str, fin_str):
    try:
        inicio = datetime.strptime(str(inicio_str), "%d/%m/%Y")
        fin = datetime.strptime(str(fin_str), "%d/%m/%Y")
        dias = (fin - inicio).days
        if dias <= 31:
            return "Mensual"
        elif dias <= 93:
            return "Trimestral"
        elif dias <= 186:
            return "Semestral"
        elif dias <= 370:
            return "Anual"
        else:
            return "Otra"
    except:
        return "No Definido"

# Identificar empresa y hoja
def identificar_empresa_y_hoja(wb):
    for hoja in wb.sheetnames:
        for celda in ["B2", "C2", "D2"]:
            valor = wb[hoja][celda].value
            if valor:
                valor_lower = str(valor).lower()
                if ("factura de crédito fiscal electrónica" in valor_lower) or ("factura de consumo electrónica" in valor_lower):
                    return "humano", hoja
    return "yunen", "Sheet1"

# Extraer datos por archivo
def extraer_datos_factura(file_path):
    wb = load_workbook(file_path, data_only=True)
    empresa, hoja_base = identificar_empresa_y_hoja(wb)
    hojas = wb.sheetnames

    datos = {
        "Cliente": "",
        "RNC Cliente": "",
        "Número Factura": "",
        "Fecha Emisión": "",
        "Dirección": "",
        "Teléfono": "",
        "Plan": "",
        "Monto": "",
        "NCF": "",
        "Vigencia": "",
        "Tipo Vigencia": "",
        "Aseguradora": ""
    }

    if empresa == "humano":
        hoja_uso = hoja_base
        sh = wb[hoja_uso]

        datos["Cliente"] = sh["B9"].value or ""
        datos["Dirección"] = sh["B9"].value or ""

        rnc_raw = f"{sh['B11'].value or ''} {sh['B10'].value or ''}"
        datos["RNC Cliente"] = re.sub(r"[^\d]", "", rnc_raw)

        datos["NCF"] = sh["E7"].value or sh["D7"].value or ""
        datos["Monto"] = sh["E20"].value or ""
        datos["Fecha Emisión"] = sh["B8"].value or ""
        datos["Número Factura"] = f"{sh['A17'].value or ''} {sh['A18'].value or ''} {sh['A19'].value or ''}".strip()

        hoja_vigencia = "Sheet2" if hoja_uso == "Sheet1" else "Sheet3"
        try:
            inicio = wb[hoja_vigencia]["B5"].value
            fin = wb[hoja_vigencia]["C5"].value
            datos["Vigencia"] = f"{inicio} - {fin}"
            datos["Tipo Vigencia"] = clasificar_vigencia(inicio, fin)
        except:
            datos["Vigencia"] = ""
            datos["Tipo Vigencia"] = "No Definido"

        hoja_plan = next((h for h in hojas if str(wb[h]["A1"].value or "").strip().lower() == "detalle de facturación"), None)
        if hoja_plan:
            datos["Plan"] = f"{wb[hoja_plan]['A8'].value or ''} {wb[hoja_plan]['A9'].value or ''}".strip()

        datos["Aseguradora"] = sh["A4"].value or "Humano"

    else:  # Yunen
        sh_base = wb[hoja_base]
        datos["NCF"] = sh_base["B8"].value or ""
        datos["Fecha Emisión"] = sh_base["A9"].value or ""

        # Número de factura en E10 a E14 del sheet base
        for fila in range(10, 15):
            valor = sh_base[f"E{fila}"].value
            if valor and re.search(r"^\d{6,}$", str(valor)):
                datos["Número Factura"] = str(valor)
                break

        # Extraer RNC sin letras
        rnc_raw = sh_base["A11"].value or ""
        datos["RNC Cliente"] = re.sub(r"[^\d]", "", str(rnc_raw))

        # Buscar monto en Sheet2, columna F para texto TOTAL FACTURADO:, monto en columna H misma fila
        monto = ""
        if "Sheet2" in wb.sheetnames:
            sh2 = wb["Sheet2"]
            for fila in range(1, sh2.max_row + 1):
                celda_f = sh2[f"F{fila}"].value
                if celda_f and isinstance(celda_f, str) and "total facturado:" in celda_f.lower():
                    celda_h = sh2[f"H{fila}"].value
                    if celda_h:
                        try:
                            monto_str = str(celda_h).replace(",", "").replace("$", "").strip()
                            monto = float(monto_str)
                        except:
                            monto = celda_h  # fallback si no se puede convertir
                    break
        datos["Monto"] = monto

        datos["Dirección"] = sh_base["B13"].value or ""
        datos["Teléfono"] = sh_base["B16"].value or ""
        datos["Plan"] = sh_base["A19"].value or ""
        datos["Cliente"] = sh_base["B12"].value or ""

        raw_vigencia = f"{sh_base['D13'].value or ''} {sh_base['E13'].value or ''} {sh_base['F13'].value or ''}".lower()
        fechas = re.findall(r"\d{2}/\d{2}/\d{4}", raw_vigencia)
        if len(fechas) >= 2:
            datos["Vigencia"] = f"{fechas[0]} - {fechas[1]}"
            datos["Tipo Vigencia"] = clasificar_vigencia(fechas[0], fechas[1])
        else:
            datos["Vigencia"] = raw_vigencia
            datos["Tipo Vigencia"] = "No Definido"

        datos["Aseguradora"] = sh_base["B1"].value or "Yunen"

    return datos


# Procesar todos los archivos
def procesar_facturas_en_carpeta():
    carpeta = "facturas"
    if not os.path.exists(carpeta):
        print("❌ Carpeta 'facturas' no encontrada.")
        return

    datos_facturas = []
    for archivo in os.listdir(carpeta):
        if archivo.endswith(".xlsx"):
            ruta = os.path.join(carpeta, archivo)
            try:
                datos = extraer_datos_factura(ruta)
                datos_facturas.append(datos)
                print(f"✅ Procesado: {archivo}")
            except Exception as e:
                print(f"❌ Error en {archivo}: {e}")

    if datos_facturas:
        df = pd.DataFrame(datos_facturas)
        df.to_excel("datos_facturas.xlsx", index=False)
        print("✅ Archivo 'datos_facturas.xlsx' generado con éxito.")
    else:
        print("⚠️ No se encontraron datos válidos para procesar.")

# Ejecutar script
if __name__ == "__main__":
    procesar_facturas_en_carpeta()

import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime

# ------------------------------------
# Función para clasificar tipo vigencia
# ------------------------------------
def clasificar_vigencia(vigencia_str):
    try:
        if pd.isna(vigencia_str):
            return "No Definido"
        partes = str(vigencia_str).split()
        # Para vigencia tipo '01/01/2023 al 31/12/2023' o solo fechas
        fechas = [p for p in partes if "/" in p]
        if len(fechas) >= 2:
            inicio = datetime.strptime(fechas[0], "%d/%m/%Y")
            fin = datetime.strptime(fechas[1], "%d/%m/%Y")
            dias = (fin - inicio).days

            if dias <= 31:
                return "Mensual"
            elif dias <= 93:
                return "Trimestral"
            elif dias <= 186:
                return "Semestral"
            elif dias <= 370:
                return "Anual"
            else:
                return "Otra"
        return "No Definido"
    except:
        return "No Definido"

# ------------------------------------
# Cargar datos de datos_facturas.xlsx
# ------------------------------------
@st.cache_data
def cargar_datos_web():
    df = pd.read_excel("datos_facturas.xlsx")
    # En caso de que falte alguna columna asegúrate que esté
    if "Aseguradora" not in df.columns:
        df["Aseguradora"] = df["Cliente"].apply(lambda x: "Humano" if "humano" in str(x).lower() else "Yunen")
    if "Tipo Vigencia" not in df.columns:
        df["Tipo Vigencia"] = df["Vigencia"].apply(clasificar_vigencia)
    return df

df_web = cargar_datos_web()

# ------------------------------------
# Streamlit Página Web para Resumen
# ------------------------------------
import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
import plotly.express as px
import os

st.set_page_config(page_title="Reporte de Facturas", layout="wide")

RUTA_ARCHIVO = "datos_facturas.xlsx"

# ------------------------
# Función: Clasificar Tipo de Vigencia
# ------------------------
def clasificar_vigencia(vigencia_str):
    try:
        partes = str(vigencia_str).replace("al", "").split()
        fechas = [p for p in partes if "/" in p]
        if len(fechas) >= 2:
            inicio = datetime.strptime(fechas[0], "%d/%m/%Y")
            fin = datetime.strptime(fechas[1], "%d/%m/%Y")
            dias = (fin - inicio).days
            if dias <= 31:
                return "Mensual"
            elif dias <= 93:
                return "Trimestral"
            elif dias <= 186:
                return "Semestral"
            elif dias <= 370:
                return "Anual"
            else:
                return "Otra"
        return "No Definido"
    except:
        return "No Definido"

# ------------------------
# Función: Cargar Datos
# ------------------------
@st.cache_data
def cargar_datos():
    if not os.path.exists(RUTA_ARCHIVO):
        return pd.DataFrame()
    df = pd.read_excel(RUTA_ARCHIVO)
    if "Aseguradora" not in df.columns:
        df["Aseguradora"] = df["Cliente"].apply(lambda x: "Humano" if "humano" in str(x).lower() else "Yunen")
    df["Tipo Vigencia"] = df["Vigencia"].apply(clasificar_vigencia)
    return df

# ------------------------
# Cargar Datos
# ------------------------
df = cargar_datos()

# ------------------------
# Tabs
# ------------------------
tabs = st.tabs(["📊 Resumen General", "👥 Clientes", "🔄 Actualizar Datos"])

# ------------------------
# TAB 1 - Resumen General
# ------------------------
with tabs[0]:
    st.header("📊 Resumen General de Facturas")

    if df.empty:
        st.warning("No hay datos disponibles. Ve al tab 'Actualizar Datos'.")
    else:
        # Resumen numérico
        total_facturas = len(df)
        total_clientes = df["Cliente"].nunique()

        facturas_por_aseguradora = df["Aseguradora"].value_counts()
        clientes_por_aseguradora = df.groupby("Aseguradora")["Cliente"].nunique()

        if "Plan" in df.columns:
            limpiar_nombre = lambda x: re.sub(r'\bplan\b', '', str(x), flags=re.IGNORECASE).strip()
            df["Plan_Limpio"] = df["Plan"].apply(limpiar_nombre)
            total_planes = df["Plan_Limpio"].nunique()
        else:
            total_planes = "N/A"

        st.markdown("---")
        st.subheader("📌 Resumen de Datos")
        col1, col2, col3 = st.columns(3)
        col1.metric("📄 Facturas Totales", f"{total_facturas:,}")
        col2.metric("👥 Clientes Únicos", f"{total_clientes:,}")
        col3.metric("📦 Planes Únicos", total_planes)

        st.markdown("### 📊 Facturas y Clientes por Aseguradora")
        resumen_df = pd.DataFrame({
            "Facturas": facturas_por_aseguradora,
            "Clientes": clientes_por_aseguradora
        }).reset_index().rename(columns={"index": "Aseguradora"})
        st.dataframe(resumen_df, use_container_width=True)

        st.markdown("---")

        # Gráfico: Cantidad de Planes por Nombre con filtro aseguradora
        st.subheader("📦 Cantidad de Planes por Nombre")

        aseguradora_seleccionada = st.selectbox(
            "Filtrar por Aseguradora:",
            options=["Todos"] + sorted(df["Aseguradora"].unique().tolist())
        )

        if aseguradora_seleccionada != "Todos":
            df_filtrado = df[df["Aseguradora"] == aseguradora_seleccionada]
        else:
            df_filtrado = df.copy()

        conteo_planes = df_filtrado["Plan_Limpio"].value_counts().reset_index()
        conteo_planes.columns = ["Plan", "Cantidad"]

        if conteo_planes.empty:
            st.info("No hay datos para mostrar con el filtro seleccionado.")
        else:
            fig_planes = px.bar(
                conteo_planes.sort_values("Cantidad", ascending=True),
                x="Cantidad",
                y="Plan",
                orientation='h',
                text="Cantidad",
                color="Cantidad",
                color_continuous_scale=px.colors.sequential.Viridis,
                title=f"Cantidad de Planes por Nombre {'- ' + aseguradora_seleccionada if aseguradora_seleccionada != 'Todos' else ''}",
                labels={"Cantidad": "Cantidad", "Plan": "Nombre del Plan"}
            )
            fig_planes.update_traces(texttemplate='%{text}', textposition='outside')
            fig_planes.update_layout(
                margin=dict(l=100, r=20, t=60, b=40),
                yaxis=dict(tickfont=dict(size=12)),
                coloraxis_showscale=False
            )
            st.plotly_chart(fig_planes, use_container_width=True)

        st.markdown("---")
        st.subheader("🍰 Distribución de Tipos de Vigencia por Aseguradora")

        col1, col2 = st.columns(2)

        # Gráfico de pastel para Humano
        df_humano = df[df["Aseguradora"].str.contains("humano", case=False, na=False)]
        counts_humano = df_humano["Tipo Vigencia"].value_counts().reset_index()
        counts_humano.columns = ["Tipo Vigencia", "Cantidad"]

        if counts_humano.empty:
            col1.info("No hay datos de tipo vigencia para Humano.")
        else:
            colores_azul_oscuro = ['#003f5c', '#2f4b7c', '#665191', '#a05195', '#d45087']
            fig_humano = px.pie(
                counts_humano,
                names="Tipo Vigencia",
                values="Cantidad",
                title="Tipo de Vigencia - Humano",
                color="Tipo Vigencia",
                color_discrete_sequence=colores_azul_oscuro,
            )
            fig_humano.update_traces(textinfo="value+label")
            fig_humano.update_layout(legend_title_text="Tipo Vigencia")
            col1.plotly_chart(fig_humano, use_container_width=True)

        # Gráfico de pastel para Yunen
        df_yunen = df[df["Aseguradora"].str.contains("yunen", case=False, na=False)]
        counts_yunen = df_yunen["Tipo Vigencia"].value_counts().reset_index()
        counts_yunen.columns = ["Tipo Vigencia", "Cantidad"]

        if counts_yunen.empty:
            col2.info("No hay datos de tipo vigencia para Yunen.")
        else:
            colores_morado_oscuro = ['#4b0082', '#551a8b', '#6a0dad', '#800080', '#9932cc']
            fig_yunen = px.pie(
                counts_yunen,
                names="Tipo Vigencia",
                values="Cantidad",
                title="Tipo de Vigencia - Yunen",
                color="Tipo Vigencia",
                color_discrete_sequence=colores_morado_oscuro,
            )
            fig_yunen.update_traces(textinfo="value+label")
            fig_yunen.update_layout(legend_title_text="Tipo Vigencia")
            col2.plotly_chart(fig_yunen, use_container_width=True)

# ------------------------
# TAB 2 - Clientes
# ------------------------
with tabs[1]:
    st.header("👥 Detalles por Cliente")

    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        clientes = df["Cliente"].dropna().unique()
        cliente_seleccionado = st.selectbox("Selecciona un cliente:", sorted(clientes))

        df_cliente = df[df["Cliente"] == cliente_seleccionado]

        if not df_cliente.empty:
            st.dataframe(df_cliente)

            fechas_finales = []
            for vig in df_cliente["Vigencia"].dropna():
                fechas = [datetime.strptime(f, "%d/%m/%Y") for f in re.findall(r"\d{2}/\d{2}/\d{4}", vig)]
                if len(fechas) == 2:
                    fechas_finales.append(fechas[1])

            if fechas_finales:
                fecha_vencimiento_max = max(fechas_finales)
                fecha_vencimiento_str = fecha_vencimiento_max.strftime("%d/%m/%Y")
            else:
                fecha_vencimiento_str = "No disponible"

            st.info(f"**Fecha de vencimiento de la última vigencia:** {fecha_vencimiento_str}")

            fig = px.bar(
                df_cliente,
                x="Número Factura",
                y="Monto",
                text="Monto",
                title="Montos por Factura",
                labels={"Monto": "RD$"},
            )
            fig.update_traces(texttemplate="%{text:.2f}", textposition="outside")
            st.plotly_chart(fig, use_container_width=True)

            cliente_limpio = re.sub(r'[\n\r\t\\/:"*?<>|]+', ' ', cliente_seleccionado).strip()

            buffer = BytesIO()
            df_cliente.to_excel(buffer, index=False)
            buffer.seek(0)
            st.download_button(
                "📥 Descargar Excel del Cliente",
                data=buffer,
                file_name=f"{cliente_limpio}_facturas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ------------------------
# TAB 3 - Actualizar Datos
# ------------------------
with tabs[2]:
    st.header("🔄 Actualizar Datos")
    st.markdown("Sube los archivos procesados previamente para refrescar el reporte.")

    uploaded = st.file_uploader("Cargar archivo de datos_facturas.xlsx", type=["xlsx"], accept_multiple_files=False)
    if uploaded:
        try:
            df_new = pd.read_excel(uploaded)
            df_new["Tipo Vigencia"] = df_new["Vigencia"].apply(clasificar_vigencia)
            df_new["Aseguradora"] = df_new["Cliente"].apply(lambda x: "Humano" if "humano" in str(x).lower() else "Yunen")
            df_new.to_excel(RUTA_ARCHIVO, index=False)
            st.success("✅ Archivo actualizado correctamente.")
        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {e}")
